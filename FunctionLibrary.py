## radius function library

import os,shutil, glob, csv, unidecode
from unidecode import unidecode
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def printme():
	print ("I am here now :-)")

def csv_to_excel(src_filepath, delimiter, delimiter_txt, dst_filepath, excel_title):
    csv_file = open(src_filepath, 'r')
    csv.register_dialect(delimiter_txt, delimiter=delimiter)
    reader = csv.reader(csv_file, dialect=delimiter_txt)

    wb = Workbook()
    ws = wb.worksheets[0]
    ws.title = excel_title

    for row_index, row in enumerate(reader):
        for column_index, cell in enumerate(row):
            column_letter = get_column_letter((column_index + 1))
            ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

    wb.save(filename = excel_path)
    print("total rows / column: ",(ws.max_row,ws.max_column))
    csv_file.close()	
	

def remove_spl_characters_from_file(src_filepath, dst_filepath):
	
	counter = 1
	if os.path.exists(dst_filepath):
	   dst_path = dst_filepath
	else:
	   dst_path = os.mkdir(dst_filepath)
	   dst_path = dst_filepath
	for filename in os.listdir(src_filepath):
		if os.path.isfile(src_filepath+filename):
		   print(str(counter)+":"+filename)
		   counter = counter+1
		   ogfilepath = src_filepath+filename
		   print("original file name: ",ogfilepath)
		   convfilepath = dst_path+filename
		   print("converted file name: ",convfilepath)
		   originalFile = open(ogfilepath,"r",encoding="latin-1") 
		   convertFile = open(convfilepath,"w",encoding="ascii") 
		   for line in originalFile:
			   line = unidecode(line)
			   line = line.replace("\x1a","")
			   convertFile.write(line)
		   convertFile.close
		   print("**File writing process completed")
		else:
		   print(filename +": is not a file")
	print("***********End of process************")
	return (originalFile,convertFile)

def convert_to_excel(src_filepath, delimiter, delimiter_txt, dst_filepath, excel_title):
	delimiter = ","
	delimiter_txt = "comma"
	csv_file = open(src_filepath, 'r')
	csv.register_dialect(delimiter_txt, delimiter=delimiter)
	reader = csv.reader(csv_file, dialect=delimiter_txt)

	wb = Workbook()
	ws = wb.worksheets[0]
	ws.title = excel_title

	for row_index, row in enumerate(reader):
		for column_index, cell in enumerate(row):
			column_letter = get_column_letter((column_index + 1))
			ws.cell('%s%s'%(column_letter, (row_index + 1))).value = cell

	wb.save(filename = dst_filepath)
	print("total rows / column: ",(ws.max_row,ws.max_column))
	csv_file.close()
	wb.close()
		#os.unlink(csv_path)
	